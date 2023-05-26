from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title  = "Data"

for row in range (1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)
        print(ws[char + str(row)].value)
wb.save(filename='Test.xlsx')